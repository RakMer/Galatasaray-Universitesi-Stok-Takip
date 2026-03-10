from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from flask_cors import CORS
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, Ekipman, EkipmanHareket, Kategori, User
from datetime import datetime
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Türkçe karakter desteği için Arial Unicode font kaydı
_ARIAL_UNICODE = '/System/Library/Fonts/Supplemental/Arial Unicode.ttf'
pdfmetrics.registerFont(TTFont('ArialUnicode', _ARIAL_UNICODE))

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///stok_takip.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# Flask-Login ayarları
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmanız gerekiyor.'


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': 'Giriş yapmanız gerekiyor.'}), 401
    return redirect(url_for('login'))

# Giriş sayfası
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.aktif and user.check_password(password):
            login_user(user)
            if request.is_json:
                return jsonify({'success': True, 'user': user.to_dict()})
            return redirect(url_for('index'))
        if request.is_json:
            return jsonify({'success': False, 'error': 'Kullanıcı adı veya şifre hatalı.'}), 401
        return render_template('login.html', error='Kullanıcı adı veya şifre hatalı.')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/api/kullanici/bilgi')
@login_required
def kullanici_bilgi():
    return jsonify({'success': True, 'user': current_user.to_dict()})


@app.route('/api/kullanicilar', methods=['GET'])
@login_required
def get_kullanicilar():
    if current_user.rol != 'admin':
        return jsonify({'success': False, 'error': 'Yetkiniz yok.'}), 403
    users = User.query.all()
    return jsonify([u.to_dict() for u in users])


@app.route('/api/kullanicilar', methods=['POST'])
@login_required
def add_kullanici():
    if current_user.rol != 'admin':
        return jsonify({'success': False, 'error': 'Yetkiniz yok.'}), 403
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    ad_soyad = data.get('ad_soyad', '').strip()
    rol = data.get('rol', 'kullanici')
    if not username or not password:
        return jsonify({'success': False, 'error': 'Kullanıcı adı ve şifre zorunludur.'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'error': 'Bu kullanıcı adı zaten kayıtlı.'}), 400
    user = User(username=username, ad_soyad=ad_soyad, rol=rol)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return jsonify({'success': True, 'user': user.to_dict()}), 201


@app.route('/api/kullanicilar/<int:user_id>', methods=['DELETE'])
@login_required
def delete_kullanici(user_id):
    if current_user.rol != 'admin':
        return jsonify({'success': False, 'error': 'Yetkiniz yok.'}), 403
    if user_id == current_user.id:
        return jsonify({'success': False, 'error': 'Kendinizi silemezsiniz.'}), 400
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True, 'message': f'{user.username} silindi.'})


@app.route('/api/sifre-degistir', methods=['POST'])
@login_required
def sifre_degistir():
    data = request.get_json()
    eski_sifre = data.get('eski_sifre', '')
    yeni_sifre = data.get('yeni_sifre', '')
    if not current_user.check_password(eski_sifre):
        return jsonify({'success': False, 'error': 'Mevcut şifre hatalı.'}), 400
    if len(yeni_sifre) < 4:
        return jsonify({'success': False, 'error': 'Yeni şifre en az 4 karakter olmalı.'}), 400
    current_user.set_password(yeni_sifre)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Şifre başarıyla değiştirildi.'})


# Ana sayfa
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/test')
def test():
    return render_template('test.html')

# Kategori endpoints
@app.route('/api/kategoriler', methods=['GET'])
@login_required
def get_kategoriler():
    """Tüm kategorileri getir"""
    kategoriler = Kategori.query.all()
    return jsonify([k.to_dict() for k in kategoriler])

@app.route('/api/kategoriler', methods=['POST'])
@login_required
def add_kategori():
    """Yeni kategori ekle"""
    data = request.json
    try:
        ad = data.get('ad', '').strip()
        if not ad:
            return jsonify({'success': False, 'error': 'Kategori adı boş olamaz.'}), 400
        
        if Kategori.query.filter_by(ad=ad).first():
            return jsonify({'success': False, 'error': f'"{ad}" kategorisi zaten mevcut.'}), 400
        
        kategori = Kategori(ad=ad, aciklama=data.get('aciklama', ''))
        db.session.add(kategori)
        db.session.commit()
        return jsonify({'success': True, 'data': kategori.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/kategoriler/<int:id>', methods=['DELETE'])
@login_required
def delete_kategori(id):
    """Kategori sil"""
    try:
        kategori = Kategori.query.get_or_404(id)
        
        # Bu kategoride ekipman var mı kontrol et
        ekipman_sayisi = Ekipman.query.filter_by(kategori=kategori.ad).count()
        if ekipman_sayisi > 0:
            return jsonify({
                'success': False, 
                'error': f'Bu kategoride {ekipman_sayisi} adet ekipman var. Önce ekipmanları silin veya başka kategoriye taşıyın.'
            }), 400
        
        kategori_adi = kategori.ad
        db.session.delete(kategori)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'"{kategori_adi}" kategorisi silindi.'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

# Ekipman endpoints
@app.route('/api/ekipman', methods=['GET'])
@login_required
def get_ekipman():
    """Tüm ekipmanları getir (filtreleme destekli)"""
    kategori = request.args.get('kategori')
    durum = request.args.get('durum')
    arama = request.args.get('arama')
    tedarikci = request.args.get('tedarikci')
    fiyat_min = request.args.get('fiyat_min', type=float)
    fiyat_max = request.args.get('fiyat_max', type=float)
    tarih_min = request.args.get('tarih_min')
    tarih_max = request.args.get('tarih_max')
    
    query = Ekipman.query
    
    if kategori:
        query = query.filter_by(kategori=kategori)
    if durum:
        query = query.filter_by(durum=durum)
    if arama:
        search_pattern = f'%{arama}%'
        query = query.filter(
            db.or_(
                Ekipman.marka.like(search_pattern),
                Ekipman.model.like(search_pattern),
                Ekipman.seri_no.like(search_pattern),
                Ekipman.barkod.like(search_pattern)
            )
        )
    if tedarikci:
        query = query.filter(Ekipman.tedarikci.like(f'%{tedarikci}%'))
    if fiyat_min is not None:
        query = query.filter(Ekipman.temin_fiyati >= fiyat_min)
    if fiyat_max is not None:
        query = query.filter(Ekipman.temin_fiyati <= fiyat_max)
    if tarih_min:
        try:
            query = query.filter(Ekipman.temin_tarihi >= datetime.fromisoformat(tarih_min))
        except ValueError:
            pass
    if tarih_max:
        try:
            query = query.filter(Ekipman.temin_tarihi <= datetime.fromisoformat(tarih_max))
        except ValueError:
            pass
    
    ekipmanlar = query.order_by(Ekipman.olusturma_tarihi.desc()).all()
    return jsonify([e.to_dict() for e in ekipmanlar])

@app.route('/api/ekipman/<int:id>', methods=['GET'])
@login_required
def get_ekipman_detay(id):
    """Belirli bir ekipmanın detaylarını getir"""
    ekipman = Ekipman.query.get_or_404(id)
    return jsonify(ekipman.to_dict())

@app.route('/api/ekipman', methods=['POST'])
@login_required
def ekipman_ekle():
    """Yeni ekipman ekle"""
    data = request.json
    
    try:
        # Kategori kontrolü - eğer veritabanında yoksa ekle
        kategori_adi = data['kategori']
        kategori = Kategori.query.filter_by(ad=kategori_adi).first()
        
        if not kategori:
            # Yeni kategori oluştur
            yeni_kategori = Kategori(ad=kategori_adi, aciklama='Kullanıcı tanımlı')
            db.session.add(yeni_kategori)
            db.session.commit()
            print(f"Yeni kategori eklendi: {kategori_adi}")
        
        # Tarih dönüşümü
        temin_tarihi = None
        if data.get('temin_tarihi'):
            temin_tarihi = datetime.fromisoformat(data['temin_tarihi'].replace('Z', '+00:00'))
        
        ekipman = Ekipman(
            kategori=data['kategori'],
            marka=data.get('marka'),
            model=data.get('model'),
            seri_no=data.get('seri_no'),
            barkod=data.get('barkod'),
            durum=data.get('durum', 'Depoda'),
            notlar=data.get('notlar'),
            temin_tarihi=temin_tarihi,
            temin_fiyati=data.get('temin_fiyati'),
            tedarikci=data.get('tedarikci')
        )
        
        db.session.add(ekipman)
        db.session.commit()
        
        return jsonify({'success': True, 'id': ekipman.id, 'data': ekipman.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/ekipman/<int:id>', methods=['PUT'])
@login_required
def ekipman_guncelle(id):
    """Ekipman bilgilerini güncelle"""
    ekipman = Ekipman.query.get_or_404(id)
    data = request.json
    
    try:
        if 'kategori' in data:
            ekipman.kategori = data['kategori']
        if 'marka' in data:
            ekipman.marka = data['marka']
        if 'model' in data:
            ekipman.model = data['model']
        if 'seri_no' in data:
            ekipman.seri_no = data['seri_no']
        if 'barkod' in data:
            ekipman.barkod = data['barkod']
        if 'durum' in data:
            ekipman.durum = data['durum']
        if 'notlar' in data:
            ekipman.notlar = data['notlar']
        if 'temin_tarihi' in data and data['temin_tarihi']:
            ekipman.temin_tarihi = datetime.fromisoformat(data['temin_tarihi'].replace('Z', '+00:00'))
        if 'temin_fiyati' in data:
            ekipman.temin_fiyati = data['temin_fiyati']
        if 'tedarikci' in data:
            ekipman.tedarikci = data['tedarikci']
        
        db.session.commit()
        return jsonify({'success': True, 'data': ekipman.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/ekipman/<int:id>', methods=['DELETE'])
@login_required
def ekipman_sil(id):
    """Ekipman sil"""
    ekipman = Ekipman.query.get_or_404(id)
    
    try:
        db.session.delete(ekipman)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Ekipman başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

# Hareket endpoints
@app.route('/api/hareket', methods=['GET'])
@login_required
def get_hareketler():
    """Tüm hareketleri getir"""
    ekipman_id = request.args.get('ekipman_id', type=int)
    
    query = EkipmanHareket.query
    if ekipman_id:
        query = query.filter_by(ekipman_id=ekipman_id)
    
    hareketler = query.order_by(EkipmanHareket.tarih.desc()).all()
    return jsonify([h.to_dict() for h in hareketler])

@app.route('/api/hareket', methods=['POST'])
@login_required
def hareket_ekle():
    """Yeni hareket ekle (giriş/çıkış)"""
    data = request.json
    
    try:
        hareket = EkipmanHareket(
            ekipman_id=data['ekipman_id'],
            hareket_tipi=data['hareket_tipi'],
            kullanici_adi=data.get('kullanici_adi'),
            kullanici_personel_no=data.get('kullanici_personel_no'),
            birim=data.get('birim'),
            lokasyon=data.get('lokasyon'),
            aciklama=data.get('aciklama'),
            teslim_alan=data.get('teslim_alan'),
            onaylayan=data.get('onaylayan')
        )
        
        # Ekipman durumunu güncelle
        ekipman = Ekipman.query.get(data['ekipman_id'])
        if hareket.hareket_tipi == 'Çıkış':
            ekipman.durum = 'Kullanımda'
        elif hareket.hareket_tipi == 'İade':
            ekipman.durum = 'Depoda'
        
        db.session.add(hareket)
        db.session.commit()
        
        return jsonify({'success': True, 'id': hareket.id, 'data': hareket.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

# İstatistik endpoints
@app.route('/api/istatistikler', methods=['GET'])
@login_required
def get_istatistikler():
    """Genel istatistikleri getir"""
    toplam_ekipman = Ekipman.query.count()
    depodaki = Ekipman.query.filter_by(durum='Depoda').count()
    kullanimda = Ekipman.query.filter_by(durum='Kullanımda').count()
    arizali = Ekipman.query.filter_by(durum='Arızalı').count()
    
    # Kategoriye göre dağılım
    kategori_dagilim = db.session.query(
        Ekipman.kategori, 
        db.func.count(Ekipman.id)
    ).group_by(Ekipman.kategori).all()
    
    return jsonify({
        'toplam_ekipman': toplam_ekipman,
        'depodaki': depodaki,
        'kullanimda': kullanimda,
        'arizali': arizali,
        'kategori_dagilim': {k: v for k, v in kategori_dagilim}
    })

@app.route('/api/export/excel')
@login_required
def export_excel():
    """Envanteri Excel olarak dışa aktar"""
    ekipmanlar = Ekipman.query.order_by(Ekipman.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Envanter"

    # Başlık satırı stilleri
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Kategori", "Marka", "Model", "Seri No", "Durum",
               "Temin Tarihi", "Temin Fiyatı (₺)", "Tedarikçi", "Notlar", "Eklenme Tarihi"]
    col_widths = [6, 15, 15, 20, 20, 12, 15, 18, 20, 30, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 25

    # Durum renk kodları
    durum_renk = {
        "Depoda":     "D4EDDA",
        "Kullanımda": "CCE5FF",
        "Arızalı":    "FFF3CD",
        "Hurda":      "F8D7DA",
    }

    for row_idx, e in enumerate(ekipmanlar, start=2):
        row_data = [
            e.id, e.kategori, e.marka or '', e.model or '',
            e.seri_no or '', e.durum,
            e.temin_tarihi.strftime('%d.%m.%Y') if e.temin_tarihi else '',
            e.temin_fiyati or '',
            e.tedarikci or '', e.notlar or '',
            e.olusturma_tarihi.strftime('%d.%m.%Y') if e.olusturma_tarihi else ''
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            if col_idx == 6:  # Durum sütunu renklendir
                renk = durum_renk.get(e.durum, "FFFFFF")
                cell.fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx != 6:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Dondur
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'envanter_{tarih}.xlsx'
    )


@app.route('/api/export/pdf')
@login_required
def export_pdf():
    """Envanteri PDF olarak dışa aktar"""
    ekipmanlar = Ekipman.query.order_by(Ekipman.id).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Başlık
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontName='ArialUnicode',
                                  fontSize=14, textColor=colors.HexColor('#8B0000'),
                                  spaceAfter=6)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontName='ArialUnicode',
                                fontSize=9, textColor=colors.grey, spaceAfter=12)

    elements.append(Paragraph('Galatasaray Üniversitesi - Bilgi İşlem', title_style))
    elements.append(Paragraph(f'Envanter Listesi  |  {datetime.now().strftime("%d.%m.%Y %H:%M")}  |  Toplam: {len(ekipmanlar)} ekipman', sub_style))

    # Tablo verisi
    table_data = [['ID', 'Kategori', 'Marka', 'Model', 'Seri No', 'Durum', 'Temin Tarihi', 'Fiyat (₺)']]
    for e in ekipmanlar:
        table_data.append([
            str(e.id),
            e.kategori or '',
            e.marka or '',
            e.model or '',
            e.seri_no or '',
            e.durum or '',
            e.temin_tarihi.strftime('%d.%m.%Y') if e.temin_tarihi else '',
            f'{e.temin_fiyati:,.0f}' if e.temin_fiyati else ''
        ])

    col_widths_pdf = [1.2*cm, 3.5*cm, 3*cm, 4*cm, 4*cm, 2.8*cm, 3*cm, 2.5*cm]

    tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#8B0000')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, -1), 'ArialUnicode'),
        ('FONTSIZE',       (0, 0), (-1, 0),  9),
        ('FONTSIZE',       (0, 1), (-1, -1), 8),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF8F8')]),
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWHEIGHT',      (0, 0), (-1, -1), 18),
    ]))
    elements.append(tbl)

    doc.build(elements)
    output.seek(0)

    tarih = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'envanter_{tarih}.pdf'
    )


@app.route('/api/zimmet-belgesi/<int:hareket_id>')
@login_required
def zimmet_belgesi(hareket_id):
    """Belirli bir hareket için zimmet belgesi PDF oluştur"""
    hareket = EkipmanHareket.query.get_or_404(hareket_id)
    ekipman = Ekipman.query.get_or_404(hareket.ekipman_id)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Stiller
    title_style = ParagraphStyle('zimmet_title', parent=styles['Title'],
                                  fontName='ArialUnicode', fontSize=18,
                                  textColor=colors.HexColor('#8B0000'),
                                  spaceAfter=4, alignment=1)
    subtitle_style = ParagraphStyle('zimmet_sub', parent=styles['Normal'],
                                     fontName='ArialUnicode', fontSize=10,
                                     textColor=colors.grey, alignment=1, spaceAfter=20)
    heading_style = ParagraphStyle('zimmet_heading', parent=styles['Normal'],
                                    fontName='ArialUnicode', fontSize=12,
                                    textColor=colors.HexColor('#8B0000'),
                                    spaceBefore=16, spaceAfter=8,
                                    borderWidth=0, leftIndent=0)
    normal_style = ParagraphStyle('zimmet_normal', parent=styles['Normal'],
                                   fontName='ArialUnicode', fontSize=10,
                                   spaceAfter=4)
    bold_style = ParagraphStyle('zimmet_bold', parent=styles['Normal'],
                                 fontName='ArialUnicode', fontSize=10,
                                 spaceAfter=4)
    footer_style = ParagraphStyle('zimmet_footer', parent=styles['Normal'],
                                   fontName='ArialUnicode', fontSize=8,
                                   textColor=colors.grey, alignment=1)

    # Başlık
    elements.append(Paragraph('GALATASARAY ÜNİVERSİTESİ', title_style))
    elements.append(Paragraph('Bilgi İşlem Daire Başkanlığı — Zimmet Belgesi', subtitle_style))

    # Belge bilgileri üst tablo
    belge_tarihi = hareket.tarih.strftime('%d.%m.%Y %H:%M') if hareket.tarih else '-'
    belge_data = [
        ['Belge No:', f'ZMT-{hareket.id:05d}', 'Tarih:', belge_tarihi],
        ['İşlem Tipi:', hareket.hareket_tipi, '', ''],
    ]
    belge_tbl = Table(belge_data, colWidths=[3*cm, 5.5*cm, 2.5*cm, 5.5*cm])
    belge_tbl.setStyle(TableStyle([
        ('FONTNAME',   (0, 0), (-1, -1), 'ArialUnicode'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('TEXTCOLOR',  (0, 0), (0, -1),  colors.HexColor('#555555')),
        ('TEXTCOLOR',  (2, 0), (2, -1),  colors.HexColor('#555555')),
        ('FONTSIZE',   (1, 0), (1, -1),  10),
        ('FONTSIZE',   (3, 0), (3, -1),  10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
        ('BOX',        (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('INNERGRID',  (0, 0), (-1, -1), 0.3, colors.HexColor('#EEEEEE')),
    ]))
    elements.append(belge_tbl)
    elements.append(Spacer(1, 12))

    # Ekipman bilgileri
    elements.append(Paragraph('EKİPMAN BİLGİLERİ', heading_style))
    ekipman_data = [
        ['Ekipman ID', str(ekipman.id)],
        ['Kategori', ekipman.kategori or '-'],
        ['Marka', ekipman.marka or '-'],
        ['Model', ekipman.model or '-'],
        ['Seri No', ekipman.seri_no or '-'],
        ['Barkod', ekipman.barkod or '-'],
        ['Durum', ekipman.durum or '-'],
        ['Temin Tarihi', ekipman.temin_tarihi.strftime('%d.%m.%Y') if ekipman.temin_tarihi else '-'],
    ]
    ek_tbl = Table(ekipman_data, colWidths=[4.5*cm, 12*cm])
    ek_tbl.setStyle(TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), 'ArialUnicode'),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('BACKGROUND',     (0, 0), (0, -1),  colors.HexColor('#F5F0F0')),
        ('TEXTCOLOR',      (0, 0), (0, -1),  colors.HexColor('#8B0000')),
        ('ALIGN',          (0, 0), (0, -1),  'LEFT'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWHEIGHT',      (0, 0), (-1, -1), 22),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
    ]))
    elements.append(ek_tbl)
    elements.append(Spacer(1, 12))

    # Zimmet bilgileri
    elements.append(Paragraph('ZİMMET BİLGİLERİ', heading_style))
    zimmet_data = [
        ['Teslim Alan Kişi', hareket.kullanici_adi or '-'],
        ['Personel No', hareket.kullanici_personel_no or '-'],
        ['Birim / Departman', hareket.birim or '-'],
        ['Lokasyon', hareket.lokasyon or '-'],
        ['Teslim Eden', hareket.teslim_alan or '-'],
        ['Onaylayan', hareket.onaylayan or '-'],
        ['Açıklama', hareket.aciklama or '-'],
    ]
    zm_tbl = Table(zimmet_data, colWidths=[4.5*cm, 12*cm])
    zm_tbl.setStyle(TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), 'ArialUnicode'),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('BACKGROUND',     (0, 0), (0, -1),  colors.HexColor('#F0F0F5')),
        ('TEXTCOLOR',      (0, 0), (0, -1),  colors.HexColor('#1a1a2e')),
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWHEIGHT',      (0, 0), (-1, -1), 22),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
    ]))
    elements.append(zm_tbl)

    # İmza alanları
    elements.append(Spacer(1, 50))
    imza_data = [
        ['Teslim Eden', 'Teslim Alan', 'Onaylayan'],
        ['', '', ''],
        ['', '', ''],
        ['', '', ''],
        ['İmza: _______________', 'İmza: _______________', 'İmza: _______________'],
        [
            hareket.teslim_alan or '(Ad Soyad)',
            hareket.kullanici_adi or '(Ad Soyad)',
            hareket.onaylayan or '(Ad Soyad)'
        ],
    ]
    imza_tbl = Table(imza_data, colWidths=[5.5*cm, 5.5*cm, 5.5*cm])
    imza_tbl.setStyle(TableStyle([
        ('FONTNAME',   (0, 0), (-1, -1), 'ArialUnicode'),
        ('FONTSIZE',   (0, 0), (-1, 0),  10),
        ('FONTSIZE',   (0, 1), (-1, -1), 9),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, 0),  colors.HexColor('#F5F5F5')),
        ('LINEBELOW',  (0, 0), (-1, 0),  0.5, colors.HexColor('#CCCCCC')),
        ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#888888')),
    ]))
    elements.append(imza_tbl)

    # Alt bilgi
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f'Bu belge Galatasaray Üniversitesi Bilgi İşlem Daire Başkanlığı tarafından {datetime.now().strftime("%d.%m.%Y %H:%M")} tarihinde oluşturulmuştur.',
        footer_style
    ))

    doc.build(elements)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'zimmet_belgesi_ZMT-{hareket.id:05d}.pdf'
    )


@app.route('/api/ekipman/toplu-durum', methods=['PUT'])
@login_required
def toplu_durum_guncelle():
    """Birden fazla ekipmanın durumunu toplu güncelle"""
    data = request.json
    ids = data.get('ids', [])
    yeni_durum = data.get('durum', '').strip()

    if not ids:
        return jsonify({'success': False, 'error': 'Hiç ekipman seçilmedi.'}), 400
    if yeni_durum not in ('Depoda', 'Kullanımda', 'Arızalı', 'Hurda'):
        return jsonify({'success': False, 'error': 'Geçersiz durum.'}), 400

    try:
        guncellenen = Ekipman.query.filter(Ekipman.id.in_(ids)).update(
            {Ekipman.durum: yeni_durum}, synchronize_session='fetch')
        db.session.commit()
        return jsonify({'success': True, 'message': f'{guncellenen} ekipman güncellendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/ekipman/toplu-sil', methods=['DELETE'])
@login_required
def toplu_sil():
    """Birden fazla ekipmanı toplu sil"""
    data = request.json
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'success': False, 'error': 'Hiç ekipman seçilmedi.'}), 400

    try:
        # Önce hareketleri sil
        EkipmanHareket.query.filter(EkipmanHareket.ekipman_id.in_(ids)).delete(synchronize_session='fetch')
        silinen = Ekipman.query.filter(Ekipman.id.in_(ids)).delete(synchronize_session='fetch')
        db.session.commit()
        return jsonify({'success': True, 'message': f'{silinen} ekipman silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Varsayılan admin kullanıcı oluştur
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', ad_soyad='Sistem Yöneticisi', rol='admin')
            admin.set_password('admin')
            db.session.add(admin)
            db.session.commit()
            print('✅ Varsayılan admin kullanıcı oluşturuldu (admin/admin)')
    app.run(debug=True, port=5001)
